import openpyxl
# create a messy dataset in excel for demonstration
def create_messy_dataset(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales Data"
    # add messy data
    sheet.append(["Date", "Product", "Quantity", "Price"])
    sheet.append(["2025-01-05", "Laptop", "3", "1200.50"])
    sheet.append(["", "Mouse", "", "22.99"])
    sheet.append(["2025-01-05", "Laptop", "3", "1200.50"])
    sheet.append(["2025-01-06", "Keyboard", "2", "None"])
    workbook.save(file_path)
    print("Messy dataset created.")

# clean the dataset
def clean_dataset(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    cleaned_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #skip rows with missiing data
        if None in row or "" in row:
            continue
        # remove duplicates
        if row not in cleaned_data:
            cleaned_data.append(row)
    #clearthe sheet and write cleaned data
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    # re-add headers
    headers = ["Date", "Product", "Quantity", "Price"]
    sheet.append(headers)
    # add cleaned data
    for row in cleaned_data:
        sheet.append(row)

    workbook.save(file_path)
    print("Dataset cleaned.")

# example usage
file_path = "messy_sales_data.xlsx"
create_messy_dataset(file_path)
clean_dataset(file_path)
    
